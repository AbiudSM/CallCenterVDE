import socket 
import threading
import os
import time, datetime
from rich.console import Console
from rich.markdown import Markdown
from openpyxl import Workbook, load_workbook

# * GLOBAL VARIABLES
console = Console()

HEADER = 64
PORT = 5050
SERVER = socket.gethostbyname(socket.gethostname())
ADDR = (SERVER, PORT)
FORMAT = 'utf-8'
DISCONNECT_MESSAGE = "!DISCONNECT"

server = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
server.bind(ADDR)


# ? Thread: connection with client
def handle_client(conn, addr):
    now_time = datetime.datetime.now().strftime('%A %H:%M:%S')
    
    # Get the client extension
    # ? if extension file is older (1 minute) delete the information
    try:
        msg_length = conn.recv(HEADER).decode(FORMAT)
        if msg_length:
            msg_length = int(msg_length)
            msg = conn.recv(msg_length).decode(FORMAT)

            console.print(f"\n[NEW CONNECTION] {msg} - {addr[0]} connected at {[now_time]}.",style='green')

            file_name = msg + '.txt'
            if file_name in os.listdir('ExtText'):
                delete_file(msg)
                time.sleep(.5)

        ext = msg 
    except Exception as e:
        ext = 'unknown'
        console.print(f"\n[NEW CONNECTION] {addr[0]} connected at {[now_time]}.",style='green')

    daily_report('INICIÓ SESIÓN', ext)

    # Send the URL to the connected extension
    connected = True
    while connected:
        try:
            # ? Get message from client (EXT)
            msg_length = conn.recv(HEADER).decode(FORMAT)
            if msg_length:
                msg_length = int(msg_length)
                msg = conn.recv(msg_length).decode(FORMAT)
                
                if msg == DISCONNECT_MESSAGE:
                    connected = False

                # ? Get file with the extension name
                for file in os.listdir('ExtText'):
                    if file.startswith(msg) and file.endswith('.txt'):
                        file_path = 'ExtText\\'+file

                        # Call information eg(entrante:8217897878)
                        with open(file_path, 'r') as f:
                            call_info = f.read().split(':')
                            call_type, call_phone = call_info

                        # Get URL
                        url_file = 'url_' + call_type + '.txt'
                        try:
                            with open('config/' + url_file) as f_url:
                                url_crm = f_url.read()
                        except Exception as e:
                            console.print(f"\n[WARNING] No se encontró el archivo url_crm.txt",style='yellow')
                        
                        # Send information
                        response_msg = url_crm + call_phone
                        conn.send(response_msg.encode(FORMAT))
                        
                        os.remove(file_path)
        except Exception as e:
            connected = False

        
    console.print(f'\n[LOGOUT] Usuario {ext} - {addr} a cerrado sesión',style='bold yellow') 
    console.print(f"\n[ACTIVE CONNECTIONS] {threading.activeCount() - 2}",style='cyan')

    daily_report('CERRÓ SESIÓN', ext)

    conn.close()


# Function to save the login and logout of the extensions every day
def daily_report(log_msg, ext):
    try:
        today, hour = datetime.datetime.now().strftime('%d/%m/%Y %H:%M:%S').split(' ')
        
        file_path = f'Reports/{ext}.xlsx'

        if os.path.exists(file_path): 
            wb = load_workbook(file_path)
            ws = wb.active
            ws.append([today,hour,log_msg])
            wb.save(file_path)

        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Reporte"

            ws.append(['Fecha','Hora','Evento'])
            ws.append([today,hour,log_msg])

            wb.save(file_path)

    except Exception as e:
        console.print(f"\n[WARNING] No se pudo crear el reporte para la extensión {ext}",style='yellow')


# Function to delete files that have more than 1 minute of modification
def delete_file(ext):
    
    file = 'ExtText/'+ext+'.txt'

    time_file = time.ctime(os.path.getmtime(file))
    time_file = datetime.datetime.strptime(time_file, '%a %b %d %H:%M:%S %Y')

    now_time = datetime.datetime.now().strftime('%a %b %d %H:%M:%S %Y')
    now_time = datetime.datetime.strptime(now_time, '%a %b %d %H:%M:%S %Y')

    delta = now_time - datetime.timedelta(minutes=1)

    if time_file < delta:
        os.remove(file)

# Run Server and start listening
def start():
    server.listen()
    console.print(f"\n[LISTENING] Server is listening on {SERVER}",style='cyan')
    while True:
        conn, addr = server.accept()
        thread = threading.Thread(target=handle_client, args=(conn, addr))
        thread.start()
        console.print(f"\n[ACTIVE CONNECTIONS] {threading.activeCount() - 1}",style='cyan')


console.print(Markdown("""# SERVIDOR CONMUTADOR"""),style='cyan')
console.print("\n[STARTING] server is starting...",style='cyan')

start()